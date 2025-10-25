"""
Visualization service for creating interactive charts from Excel files
Generates visualizations by reading data from Excel sheets
"""
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
from typing import Dict, Any, List, Optional
import os
import json
import openpyxl


class VisualizationService:
    """Service for generating interactive financial visualizations from Excel files"""
    
    def __init__(self):
        self.output_dir = "outputs/visualizations"
        os.makedirs(self.output_dir, exist_ok=True)
    
    def _load_excel_workbook(self, excel_path: str):
        """Load Excel workbook"""
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        return openpyxl.load_workbook(excel_path)
    
    def _find_cell_value(self, sheet, search_text: str, search_col: int = 1, return_col: int = 2, max_row: int = 100) -> Optional[float]:
        """
        Search for a label in a column and return the corresponding value from another column
        
        Args:
            sheet: Excel worksheet
            search_text: Text to search for in cells
            search_col: Column to search in (1-indexed)
            return_col: Column to return value from (1-indexed)
            max_row: Maximum row to search
            
        Returns:
            Float value if found, None otherwise
        """
        for row in sheet.iter_rows(min_row=1, max_row=max_row):
            if row[search_col - 1].value and search_text.lower() in str(row[search_col - 1].value).lower():
                try:
                    return float(row[return_col - 1].value) if row[return_col - 1].value else None
                except (ValueError, TypeError):
                    continue
        return None
    
    def _extract_table_data(self, sheet, start_row: int, label_col: int = 1, 
                           data_cols: List[int] = [2], max_rows: int = 50) -> List[Dict]:
        """
        Extract table data from a sheet
        
        Args:
            sheet: Excel worksheet
            start_row: Row to start reading from
            label_col: Column containing labels
            data_cols: Columns containing data values
            max_rows: Maximum rows to read
            
        Returns:
            List of dictionaries with labels and values
        """
        data = []
        for row in sheet.iter_rows(min_row=start_row, max_row=start_row + max_rows):
            if row[label_col - 1].value:
                label = str(row[label_col - 1].value).strip()
                # Skip section headers (all caps or ends with colon)
                if label.isupper() or label.endswith(':') or not any(col < len(row) and row[col - 1].value for col in data_cols):
                    continue
                
                try:
                    row_data = {'label': label}
                    for i, col in enumerate(data_cols):
                        if col - 1 < len(row):
                            value = row[col - 1].value
                            row_data[f'value_{i}'] = float(value) if value is not None else None
                    data.append(row_data)
                except (ValueError, TypeError):
                    continue
        return data
    
    def generate_all_visualizations_from_excel(self, excel_path: str, report_id: int) -> Dict[str, str]:
        """
        Generate all visualizations from Excel file
        
        Args:
            excel_path: Path to the Excel file
            report_id: ID of the report
            
        Returns:
            Dictionary mapping chart names to file paths
        """
        visualizations = {}
        
        try:
            print(f"📊 Loading Excel file: {excel_path}")
            wb = self._load_excel_workbook(excel_path)
            print(f"✅ Loaded workbook with {len(wb.sheetnames)} sheets")
            
            # 1. Revenue Comparison Chart
            print("📈 Creating Revenue Comparison...")
            revenue_path = self.create_revenue_comparison_from_excel(wb, report_id)
            if revenue_path:
                visualizations['revenue_comparison'] = revenue_path
                print(f"   ✅ Created: {revenue_path}")
            
            # 2. Financial Ratios Dashboard
            print("📊 Creating Ratios Dashboard...")
            ratios_path = self.create_ratios_dashboard_from_excel(wb, report_id)
            if ratios_path:
                visualizations['ratios_dashboard'] = ratios_path
                print(f"   ✅ Created: {ratios_path}")
            
            # 3. Balance Sheet Visualization
            print("📊 Creating Balance Sheet Chart...")
            balance_path = self.create_balance_sheet_from_excel(wb, report_id)
            if balance_path:
                visualizations['balance_sheet'] = balance_path
                print(f"   ✅ Created: {balance_path}")
            
            # 4. Cash Flow Waterfall
            print("💰 Creating Cash Flow Waterfall...")
            cash_flow_path = self.create_cash_flow_waterfall_from_excel(wb, report_id)
            if cash_flow_path:
                visualizations['cash_flow_waterfall'] = cash_flow_path
                print(f"   ✅ Created: {cash_flow_path}")
            
            # 5. Segment Analysis
            print("🔍 Creating Segment Analysis...")
            segment_path = self.create_segment_analysis_from_excel(wb, report_id)
            if segment_path:
                visualizations['segment_analysis'] = segment_path
                print(f"   ✅ Created: {segment_path}")
            
            # 6. Geographic Breakdown
            print("🌍 Creating Geographic Analysis...")
            geo_path = self.create_geographic_analysis_from_excel(wb, report_id)
            if geo_path:
                visualizations['geographic_analysis'] = geo_path
                print(f"   ✅ Created: {geo_path}")
            
            wb.close()
            print(f"✅ All visualizations generated successfully!")
            return visualizations
            
        except Exception as e:
            print(f"❌ Error generating visualizations: {str(e)}")
            import traceback
            traceback.print_exc()
            return visualizations
    
    def create_revenue_comparison_from_excel(self, wb, report_id: int) -> Optional[str]:
        """Create revenue & net income comparison chart from Excel"""
        try:
            if "Executive Summary" not in wb.sheetnames:
                return None
            
            summary_sheet = wb["Executive Summary"]
            
            # Find Revenue and Net Income rows
            revenue_current = self._find_cell_value(summary_sheet, "Revenue", 1, 2)
            revenue_previous = self._find_cell_value(summary_sheet, "Revenue", 1, 3)
            net_income_current = self._find_cell_value(summary_sheet, "Net Income", 1, 2)
            net_income_previous = self._find_cell_value(summary_sheet, "Net Income", 1, 3)
            
            if not revenue_current or not revenue_previous:
                return None
            
            fig = go.Figure()
            
            # Revenue bars
            fig.add_trace(go.Bar(
                name='Revenue',
                x=['Previous Year', 'Current Year'],
                y=[revenue_previous, revenue_current],
                marker_color='rgb(26, 118, 255)',
                text=[f'${revenue_previous:,.0f}M', f'${revenue_current:,.0f}M'],
                textposition='auto'
            ))
            
            # Net Income bars (if available)
            if net_income_current and net_income_previous:
                fig.add_trace(go.Bar(
                    name='Net Income',
                    x=['Previous Year', 'Current Year'],
                    y=[net_income_previous, net_income_current],
                    marker_color='rgb(55, 83, 109)',
                    text=[f'${net_income_previous:,.0f}M', f'${net_income_current:,.0f}M'],
                    textposition='auto'
                ))
            
            fig.update_layout(
                title='Revenue & Net Income Comparison',
                xaxis_title='Period',
                yaxis_title='Amount (Millions)',
                barmode='group',
                template='plotly_white',
                height=500
            )
            
            output_path = os.path.join(self.output_dir, f"revenue_comparison_{report_id}.html")
            fig.write_html(output_path)
            return output_path
            
        except Exception as e:
            print(f"Error creating revenue comparison: {str(e)}")
            return None
    
    def create_ratios_dashboard_from_excel(self, wb, report_id: int) -> Optional[str]:
        """Create financial ratios dashboard from Excel"""
        try:
            if "Financial Ratios" not in wb.sheetnames:
                return None
            
            ratios_sheet = wb["Financial Ratios"]
            
            # Extract key ratios
            roe = self._find_cell_value(ratios_sheet, "Return on Equity", 1, 2)
            current_ratio = self._find_cell_value(ratios_sheet, "Current Ratio", 1, 2)
            debt_to_equity = self._find_cell_value(ratios_sheet, "Debt to Equity", 1, 2)
            revenue_growth = self._find_cell_value(ratios_sheet, "Revenue Growth", 1, 2)
            
            # Create subplots with gauges
            fig = make_subplots(
                rows=2, cols=2,
                subplot_titles=('Profitability (ROE)', 'Liquidity', 'Leverage', 'Growth'),
                specs=[[{'type': 'indicator'}, {'type': 'indicator'}],
                       [{'type': 'indicator'}, {'type': 'indicator'}]]
            )
            
            # ROE Gauge
            if roe:
                roe_val = roe if roe > 1 else roe * 100  # Convert if decimal
                fig.add_trace(go.Indicator(
                    mode="gauge+number+delta",
                    value=roe_val,
                    title={'text': "ROE (%)"},
                    gauge={'axis': {'range': [0, 100]},
                           'bar': {'color': "darkblue"},
                           'steps': [{'range': [0, 10], 'color': "lightgray"},
                                   {'range': [10, 20], 'color': "gray"}],
                           'threshold': {'line': {'color': "red", 'width': 4}, 
                                       'thickness': 0.75, 'value': 15}}
                ), row=1, col=1)
            
            # Current Ratio Gauge
            if current_ratio:
                fig.add_trace(go.Indicator(
                    mode="gauge+number",
                    value=current_ratio,
                    title={'text': "Current Ratio"},
                    gauge={'axis': {'range': [0, 3]},
                           'bar': {'color': "green"},
                           'steps': [{'range': [0, 1], 'color': "lightcoral"},
                                   {'range': [1, 2], 'color': "lightyellow"}],
                           'threshold': {'line': {'color': "darkgreen", 'width': 4},
                                       'thickness': 0.75, 'value': 1.5}}
                ), row=1, col=2)
            
            # Debt to Equity Gauge
            if debt_to_equity:
                fig.add_trace(go.Indicator(
                    mode="gauge+number",
                    value=debt_to_equity,
                    title={'text': "Debt/Equity"},
                    gauge={'axis': {'range': [0, 2]},
                           'bar': {'color': "orange"},
                           'steps': [{'range': [0, 0.5], 'color': "lightgreen"},
                                   {'range': [0.5, 1], 'color': "lightyellow"}],
                           'threshold': {'line': {'color': "red", 'width': 4},
                                       'thickness': 0.75, 'value': 1}}
                ), row=2, col=1)
            
            # Revenue Growth
            if revenue_growth:
                growth_val = revenue_growth if revenue_growth > 1 else revenue_growth * 100
                fig.add_trace(go.Indicator(
                    mode="number+delta",
                    value=growth_val,
                    title={'text': "Revenue Growth (%)"},
                    number={'suffix': "%"},
                    delta={'reference': 10, 'relative': False}
                ), row=2, col=2)
            
            fig.update_layout(
                height=600,
                title_text="Key Financial Ratios Dashboard",
                template='plotly_white'
            )
            
            output_path = os.path.join(self.output_dir, f"ratios_dashboard_{report_id}.html")
            fig.write_html(output_path)
            return output_path
            
        except Exception as e:
            print(f"Error creating ratios dashboard: {str(e)}")
            return None
    
    def create_balance_sheet_from_excel(self, wb, report_id: int) -> Optional[str]:
        """Create balance sheet visualization from Excel"""
        try:
            if "Balance Sheet" not in wb.sheetnames:
                return None
            
            balance_sheet = wb["Balance Sheet"]
            
            # Extract balance sheet items
            total_current_assets = self._find_cell_value(balance_sheet, "Total Current Assets", 1, 2)
            total_non_current_assets = self._find_cell_value(balance_sheet, "Total Non-Current Assets", 1, 2)
            total_current_liabilities = self._find_cell_value(balance_sheet, "Total Current Liabilities", 1, 2)
            total_non_current_liabilities = self._find_cell_value(balance_sheet, "Total Non-Current Liabilities", 1, 2)
            total_equity = self._find_cell_value(balance_sheet, "TOTAL SHAREHOLDERS' EQUITY", 1, 2)
            
            if not (total_current_assets and total_non_current_assets):
                return None
            
            fig = go.Figure()
            
            # Assets stack
            fig.add_trace(go.Bar(
                name='Current Assets',
                x=['Assets'],
                y=[total_current_assets],
                marker_color='lightblue'
            ))
            fig.add_trace(go.Bar(
                name='Non-Current Assets',
                x=['Assets'],
                y=[total_non_current_assets],
                marker_color='darkblue'
            ))
            
            # Liabilities & Equity stack
            if total_current_liabilities:
                fig.add_trace(go.Bar(
                    name='Current Liabilities',
                    x=['Liabilities & Equity'],
                    y=[total_current_liabilities],
                    marker_color='lightcoral'
                ))
            if total_non_current_liabilities:
                fig.add_trace(go.Bar(
                    name='Non-Current Liabilities',
                    x=['Liabilities & Equity'],
                    y=[total_non_current_liabilities],
                    marker_color='darkred'
                ))
            if total_equity:
                fig.add_trace(go.Bar(
                    name='Shareholders Equity',
                    x=['Liabilities & Equity'],
                    y=[total_equity],
                    marker_color='lightgreen'
                ))
            
            fig.update_layout(
                title='Balance Sheet Composition',
                yaxis_title='Amount (Millions)',
                barmode='stack',
                template='plotly_white',
                height=600
            )
            
            output_path = os.path.join(self.output_dir, f"balance_sheet_{report_id}.html")
            fig.write_html(output_path)
            return output_path
            
        except Exception as e:
            print(f"Error creating balance sheet: {str(e)}")
            return None
    
    def create_cash_flow_waterfall_from_excel(self, wb, report_id: int) -> Optional[str]:
        """Create cash flow waterfall chart from Excel"""
        try:
            if "Cash Flow" not in wb.sheetnames:
                return None
            
            cash_flow_sheet = wb["Cash Flow"]
            
            # Extract cash flow items
            operating_cf = self._find_cell_value(cash_flow_sheet, "Net Cash from Operating Activities", 1, 2)
            investing_cf = self._find_cell_value(cash_flow_sheet, "Net Cash from Investing Activities", 1, 2)
            financing_cf = self._find_cell_value(cash_flow_sheet, "Net Cash from Financing Activities", 1, 2)
            net_change = self._find_cell_value(cash_flow_sheet, "Net Change in Cash", 1, 2)
            
            if not (operating_cf and investing_cf and financing_cf):
                return None
            
            fig = go.Figure(go.Waterfall(
                name="Cash Flow",
                orientation="v",
                measure=["relative", "relative", "relative", "total"],
                x=["Operating Activities", "Investing Activities", "Financing Activities", "Net Change"],
                textposition="outside",
                text=[f"${operating_cf:,.0f}M", f"${investing_cf:,.0f}M", 
                      f"${financing_cf:,.0f}M", f"${net_change:,.0f}M"],
                y=[operating_cf, investing_cf, financing_cf, net_change],
                connector={"line": {"color": "rgb(63, 63, 63)"}},
                decreasing={"marker": {"color": "red"}},
                increasing={"marker": {"color": "green"}},
                totals={"marker": {"color": "blue"}}
            ))
            
            fig.update_layout(
                title="Cash Flow Waterfall Chart",
                yaxis_title="Amount (Millions)",
                template='plotly_white',
                height=600
            )
            
            output_path = os.path.join(self.output_dir, f"cash_flow_waterfall_{report_id}.html")
            fig.write_html(output_path)
            return output_path
            
        except Exception as e:
            print(f"Error creating cash flow waterfall: {str(e)}")
            return None
    
    def create_segment_analysis_from_excel(self, wb, report_id: int) -> Optional[str]:
        """Create segment analysis chart from Excel"""
        try:
            if "Segment Analysis" not in wb.sheetnames:
                return None
            
            segment_sheet = wb["Segment Analysis"]
            
            # Extract segment data
            segments_data = self._extract_table_data(segment_sheet, start_row=4, label_col=1, data_cols=[2, 5])
            
            if not segments_data:
                return None
            
            segments = [d['label'] for d in segments_data]
            revenues = [d.get('value_0', 0) for d in segments_data]
            margins = [d.get('value_1', 0) for d in segments_data]
            
            # Convert margins to percentage if needed
            margins = [m * 100 if m and m < 1 else m for m in margins]
            
            # Create subplot with bar and line
            fig = make_subplots(specs=[[{"secondary_y": True}]])
            
            fig.add_trace(
                go.Bar(name="Revenue", x=segments, y=revenues, marker_color='skyblue'),
                secondary_y=False
            )
            
            fig.add_trace(
                go.Scatter(name="Margin %", x=segments, y=margins, mode='lines+markers',
                          marker=dict(size=10, color='red'), line=dict(color='red', width=3)),
                secondary_y=True
            )
            
            fig.update_xaxes(title_text="Business Segment")
            fig.update_yaxes(title_text="Revenue (Millions)", secondary_y=False)
            fig.update_yaxes(title_text="Margin (%)", secondary_y=True)
            
            fig.update_layout(
                title="Business Segment Performance",
                template='plotly_white',
                height=600
            )
            
            output_path = os.path.join(self.output_dir, f"segment_analysis_{report_id}.html")
            fig.write_html(output_path)
            return output_path
            
        except Exception as e:
            print(f"Error creating segment analysis: {str(e)}")
            return None
    
    def create_geographic_analysis_from_excel(self, wb, report_id: int) -> Optional[str]:
        """Create geographic analysis chart from Excel"""
        try:
            if "Geographic Analysis" not in wb.sheetnames:
                return None
            
            geo_sheet = wb["Geographic Analysis"]
            
            # Extract geographic data
            geo_data = self._extract_table_data(geo_sheet, start_row=4, label_col=1, data_cols=[2])
            
            if not geo_data:
                return None
            
            regions = [d['label'] for d in geo_data]
            revenues = [d.get('value_0', 0) for d in geo_data]
            
            fig = go.Figure(data=[
                go.Pie(labels=regions, values=revenues, hole=0.4)
            ])
            
            fig.update_layout(
                title="Revenue by Geographic Region",
                template='plotly_white',
                height=600
            )
            
            output_path = os.path.join(self.output_dir, f"geographic_analysis_{report_id}.html")
            fig.write_html(output_path)
            return output_path
            
        except Exception as e:
            print(f"Error creating geographic analysis: {str(e)}")
            return None
    
    # Keep old methods for backward compatibility
    def create_revenue_comparison(self, data: Dict[str, Any], analysis_id: int) -> Optional[str]:
        """
        Legacy method - Create year-over-year revenue comparison chart
        """
        try:
            metrics = data.get('financial_metrics', {})
            
            # Extract revenue data (current and previous year if available)
            current_revenue = metrics.get('total_revenue', 0)
            previous_revenue = metrics.get('previous_revenue', current_revenue * 0.9)  # Fallback
            
            # Calculate growth
            growth = ((current_revenue - previous_revenue) / previous_revenue * 100) if previous_revenue else 0
            
            fig = go.Figure()
            
            # Add bars
            fig.add_trace(go.Bar(
                x=['Previous Year', 'Current Year'],
                y=[previous_revenue, current_revenue],
                text=[f'${previous_revenue:,.0f}M', f'${current_revenue:,.0f}M'],
                textposition='auto',
                marker_color=['#636EFA', '#00CC96'],
                name='Revenue'
            ))
            
            # Add growth annotation
            fig.add_annotation(
                x=1,
                y=current_revenue,
                text=f"Growth: {growth:+.1f}%",
                showarrow=True,
                arrowhead=2,
                arrowsize=1,
                arrowwidth=2,
                arrowcolor="#00CC96",
                ax=-50,
                ay=-40
            )
            
            fig.update_layout(
                title="Revenue Comparison (Year-over-Year)",
                xaxis_title="Period",
                yaxis_title="Revenue ($ Millions)",
                template="plotly_white",
                height=500,
                showlegend=False
            )
            
            # Save
            filepath = os.path.join(self.output_dir, f"revenue_comparison_{analysis_id}.html")
            fig.write_html(filepath)
            return filepath
            
        except Exception as e:
            print(f"Error creating revenue comparison: {str(e)}")
            return None
    
    def create_metrics_dashboard(self, data: Dict[str, Any], analysis_id: int) -> Optional[str]:
        """
        Create dashboard with key financial metrics as gauges
        """
        try:
            kpis = data.get('kpis', {})
            
            # Extract KPIs
            eps = float(kpis.get('eps', 0))
            pe_ratio = float(kpis.get('pe_ratio', 0))
            roe = float(kpis.get('roe', 0))
            debt_to_equity = float(kpis.get('debt_to_equity', 0))
            
            # Create subplots with gauges
            from plotly.subplots import make_subplots
            
            fig = make_subplots(
                rows=2, cols=2,
                specs=[[{'type': 'indicator'}, {'type': 'indicator'}],
                       [{'type': 'indicator'}, {'type': 'indicator'}]],
                subplot_titles=('Earnings Per Share', 'P/E Ratio', 'Return on Equity', 'Debt-to-Equity')
            )
            
            # EPS Gauge
            fig.add_trace(go.Indicator(
                mode="gauge+number+delta",
                value=eps,
                title={'text': "EPS ($)"},
                delta={'reference': eps * 0.9},
                gauge={'axis': {'range': [None, eps * 1.5]},
                       'bar': {'color': "#636EFA"},
                       'steps': [
                           {'range': [0, eps * 0.7], 'color': "lightgray"},
                           {'range': [eps * 0.7, eps * 1.2], 'color': "lightblue"}],
                       'threshold': {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': eps}}
            ), row=1, col=1)
            
            # P/E Ratio Gauge
            fig.add_trace(go.Indicator(
                mode="gauge+number",
                value=pe_ratio,
                title={'text': "P/E Ratio"},
                gauge={'axis': {'range': [0, 50]},
                       'bar': {'color': "#00CC96"},
                       'steps': [
                           {'range': [0, 15], 'color': "lightgreen"},
                           {'range': [15, 25], 'color': "yellow"},
                           {'range': [25, 50], 'color': "lightcoral"}]}
            ), row=1, col=2)
            
            # ROE Gauge
            fig.add_trace(go.Indicator(
                mode="gauge+number",
                value=roe,
                title={'text': "ROE (%)"},
                gauge={'axis': {'range': [0, 30]},
                       'bar': {'color': "#AB63FA"},
                       'steps': [
                           {'range': [0, 10], 'color': "lightgray"},
                           {'range': [10, 20], 'color': "lightblue"},
                           {'range': [20, 30], 'color': "lightgreen"}]}
            ), row=2, col=1)
            
            # Debt-to-Equity Gauge
            fig.add_trace(go.Indicator(
                mode="gauge+number",
                value=debt_to_equity,
                title={'text': "Debt/Equity"},
                gauge={'axis': {'range': [0, 3]},
                       'bar': {'color': "#FFA15A"},
                       'steps': [
                           {'range': [0, 1], 'color': "lightgreen"},
                           {'range': [1, 2], 'color': "yellow"},
                           {'range': [2, 3], 'color': "lightcoral"}]}
            ), row=2, col=2)
            
            fig.update_layout(
                title="Key Financial Metrics Dashboard",
                height=600,
                template="plotly_white"
            )
            
            # Save
            filepath = os.path.join(self.output_dir, f"metrics_dashboard_{analysis_id}.html")
            fig.write_html(filepath)
            return filepath
            
        except Exception as e:
            print(f"Error creating metrics dashboard: {str(e)}")
            return None
    
    def create_cash_flow_waterfall(self, data: Dict[str, Any], analysis_id: int) -> Optional[str]:
        """
        Create waterfall chart for cash flow analysis
        """
        try:
            metrics = data.get('financial_metrics', {})
            
            # Extract cash flow components
            operating_cf = metrics.get('operating_cash_flow', 0)
            investing_cf = metrics.get('investing_cash_flow', 0)
            financing_cf = metrics.get('financing_cash_flow', 0)
            
            # Create waterfall data
            x = ['Operating Activities', 'Investing Activities', 'Financing Activities', 'Net Cash Flow']
            y = [operating_cf, investing_cf, financing_cf, operating_cf + investing_cf + financing_cf]
            
            fig = go.Figure(go.Waterfall(
                x=x,
                y=y,
                measure=['relative', 'relative', 'relative', 'total'],
                text=[f'${v:,.0f}M' for v in y],
                textposition='outside',
                connector={'line': {'color': 'rgb(63, 63, 63)'}},
                decreasing={'marker': {'color': '#EF553B'}},
                increasing={'marker': {'color': '#00CC96'}},
                totals={'marker': {'color': '#636EFA'}}
            ))
            
            fig.update_layout(
                title="Cash Flow Waterfall Analysis",
                xaxis_title="Cash Flow Category",
                yaxis_title="Amount ($ Millions)",
                template="plotly_white",
                height=500,
                showlegend=False
            )
            
            # Save
            filepath = os.path.join(self.output_dir, f"cash_flow_waterfall_{analysis_id}.html")
            fig.write_html(filepath)
            return filepath
            
        except Exception as e:
            print(f"Error creating cash flow waterfall: {str(e)}")
            return None
    
    def create_segment_pie_chart(self, data: Dict[str, Any], analysis_id: int) -> Optional[str]:
        """
        Create pie chart for business segment revenue distribution
        """
        try:
            segments = data.get('segments', [])
            
            if not segments:
                return None
            
            # Extract segment data
            labels = [seg.get('name', 'Unknown') for seg in segments]
            values = [seg.get('revenue', 0) for seg in segments]
            
            fig = go.Figure(data=[go.Pie(
                labels=labels,
                values=values,
                hole=0.4,  # Donut chart
                textinfo='label+percent',
                textposition='auto',
                marker=dict(colors=px.colors.qualitative.Set3)
            )])
            
            fig.update_layout(
                title="Revenue Distribution by Business Segment",
                template="plotly_white",
                height=500,
                showlegend=True
            )
            
            # Save
            filepath = os.path.join(self.output_dir, f"segment_pie_{analysis_id}.html")
            fig.write_html(filepath)
            return filepath
            
        except Exception as e:
            print(f"Error creating segment pie chart: {str(e)}")
            return None
    
    def create_geographic_bar_chart(self, data: Dict[str, Any], analysis_id: int) -> Optional[str]:
        """
        Create bar chart for geographic revenue distribution
        """
        try:
            geographic = data.get('geographic', [])
            
            if not geographic:
                return None
            
            # Extract geographic data
            regions = [geo.get('region', 'Unknown') for geo in geographic]
            revenues = [geo.get('revenue', 0) for geo in geographic]
            
            # Sort by revenue descending
            sorted_data = sorted(zip(regions, revenues), key=lambda x: x[1], reverse=True)
            regions, revenues = zip(*sorted_data) if sorted_data else ([], [])
            
            fig = go.Figure()
            
            fig.add_trace(go.Bar(
                x=list(regions),
                y=list(revenues),
                text=[f'${v:,.0f}M' for v in revenues],
                textposition='auto',
                marker_color=px.colors.qualitative.Plotly
            ))
            
            fig.update_layout(
                title="Revenue by Geographic Region",
                xaxis_title="Region",
                yaxis_title="Revenue ($ Millions)",
                template="plotly_white",
                height=500,
                showlegend=False
            )
            
            # Save
            filepath = os.path.join(self.output_dir, f"geographic_bar_{analysis_id}.html")
            fig.write_html(filepath)
            return filepath
            
        except Exception as e:
            print(f"Error creating geographic bar chart: {str(e)}")
            return None
    
    def create_balance_sheet_chart(self, data: Dict[str, Any], analysis_id: int) -> Optional[str]:
        """
        Create stacked bar chart for balance sheet composition
        """
        try:
            metrics = data.get('financial_metrics', {})
            
            # Extract balance sheet components
            total_assets = metrics.get('total_assets', 0)
            total_liabilities = metrics.get('total_liabilities', 0)
            shareholders_equity = metrics.get('shareholders_equity', 0)
            
            fig = go.Figure()
            
            # Assets bar
            fig.add_trace(go.Bar(
                x=['Balance Sheet'],
                y=[total_assets],
                name='Total Assets',
                text=f'${total_assets:,.0f}M',
                textposition='inside',
                marker_color='#636EFA'
            ))
            
            # Liabilities bar
            fig.add_trace(go.Bar(
                x=['Liabilities & Equity'],
                y=[total_liabilities],
                name='Liabilities',
                text=f'${total_liabilities:,.0f}M',
                textposition='inside',
                marker_color='#EF553B'
            ))
            
            # Equity bar (stacked on liabilities)
            fig.add_trace(go.Bar(
                x=['Liabilities & Equity'],
                y=[shareholders_equity],
                name="Shareholders' Equity",
                text=f'${shareholders_equity:,.0f}M',
                textposition='inside',
                marker_color='#00CC96'
            ))
            
            fig.update_layout(
                title="Balance Sheet Composition",
                xaxis_title="",
                yaxis_title="Amount ($ Millions)",
                barmode='stack',
                template="plotly_white",
                height=500,
                showlegend=True
            )
            
            # Save
            filepath = os.path.join(self.output_dir, f"balance_sheet_{analysis_id}.html")
            fig.write_html(filepath)
            return filepath
            
        except Exception as e:
            print(f"Error creating balance sheet chart: {str(e)}")
            return None
