"""
Custom Visualization Service - Let users create their own charts
"""
import plotly.graph_objects as go
import plotly.express as px
from typing import Dict, Any, List, Optional
import os
import openpyxl
import json


class CustomVisualizationService:
    """Service for creating custom user-defined visualizations"""
    
    def __init__(self):
        self.output_dir = "outputs/visualizations"
        os.makedirs(self.output_dir, exist_ok=True)
    
    def _detect_tables_in_sheet(self, sheet) -> List[Dict[str, Any]]:
        """
        Smart table detection: finds actual data tables in a sheet
        Returns list of tables with their headers and data
        """
        tables = []
        max_rows = min(sheet.max_row, 200)  # Limit to 200 rows
        max_cols = min(sheet.max_column, 20)  # Limit to 20 columns
        
        i = 1
        while i <= max_rows:
            row_cells = [sheet.cell(i, j) for j in range(1, max_cols + 1)]
            
            # Skip completely empty rows
            if all(cell.value is None for cell in row_cells):
                i += 1
                continue
            
            # Skip title rows (merged cells, background fill, or mostly empty)
            non_empty = [cell for cell in row_cells if cell.value is not None]
            if len(non_empty) <= 2:  # Title rows usually have 1-2 values
                i += 1
                continue
            
            # Check if this could be a header row (next row has numeric data)
            if i < max_rows:
                potential_headers = [str(cell.value).strip() if cell.value else f"Column{j}" 
                                   for j, cell in enumerate(row_cells, 1) if cell.value is not None]
                
                # Look ahead to see if next rows have numeric data
                next_row_cells = [sheet.cell(i+1, j) for j in range(1, max_cols + 1)]
                has_numbers = sum(1 for cell in next_row_cells if isinstance(cell.value, (int, float)))
                
                # If next row has at least 2 numbers, this is likely a header
                if has_numbers >= 2 and len(potential_headers) >= 2:
                    table_data = self._extract_table_data(sheet, i, max_rows, max_cols)
                    if table_data['rows']:
                        tables.append(table_data)
                        i = table_data['end_row'] + 1
                        continue
            
            i += 1
        
        return tables
    
    def _extract_table_data(self, sheet, header_row: int, max_rows: int, max_cols: int) -> Dict[str, Any]:
        """Extract a table starting from header_row"""
        # Get headers
        header_cells = [sheet.cell(header_row, j) for j in range(1, max_cols + 1)]
        headers = []
        header_indices = []
        
        for j, cell in enumerate(header_cells, 1):
            if cell.value is not None:
                header_text = str(cell.value).strip()
                # Skip if header is too long (likely not a real header)
                if len(header_text) < 50:
                    headers.append(header_text)
                    header_indices.append(j)
        
        if len(headers) < 2:
            return {'rows': [], 'end_row': header_row}
        
        # Extract data rows
        rows = []
        current_row = header_row + 1
        consecutive_empty = 0
        
        while current_row <= max_rows and consecutive_empty < 3:
            row_data = {}
            has_data = False
            
            for idx, col in enumerate(header_indices):
                cell = sheet.cell(current_row, col)
                value = cell.value
                
                if value is not None:
                    has_data = True
                    # Convert to appropriate type
                    if isinstance(value, (int, float)):
                        row_data[headers[idx]] = float(value)
                    else:
                        row_data[headers[idx]] = str(value)
            
            if has_data:
                rows.append(row_data)
                consecutive_empty = 0
            else:
                consecutive_empty += 1
            
            current_row += 1
        
        return {
            'headers': headers,
            'rows': rows,
            'start_row': header_row,
            'end_row': current_row - consecutive_empty,
            'row_count': len(rows)
        }
    
    def get_excel_data_structure(self, excel_path: str) -> Dict[str, Any]:
        """
        Extract available data structure from Excel file with smart table detection
        
        Returns:
        {
            "sheets": [
                {
                    "name": "Income Statement",
                    "tables": [
                        {
                            "table_id": 0,
                            "columns": ["Line Item", "FY 2024", "FY 2023"],
                            "row_count": 25,
                            "preview": [...]
                        }
                    ]
                }
            ]
        }
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        wb = openpyxl.load_workbook(excel_path)
        sheets_data = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Detect all tables in this sheet
            tables = self._detect_tables_in_sheet(sheet)
            
            if tables:
                # Format tables for frontend
                formatted_tables = []
                for idx, table in enumerate(tables):
                    formatted_tables.append({
                        "table_id": idx,
                        "columns": table['headers'],
                        "row_count": table['row_count'],
                        "preview": table['rows'][:10],  # First 10 rows
                        "start_row": table['start_row'],
                        "end_row": table['end_row']
                    })
                
                sheets_data.append({
                    "name": sheet_name,
                    "tables": formatted_tables,
                    "table_count": len(formatted_tables)
                })
        
        return {
            "sheets": sheets_data,
            "total_sheets": len(sheets_data)
        }
    
    def create_custom_chart(
        self,
        excel_path: str,
        report_id: int,
        chart_config: Dict[str, Any]
    ) -> str:
        """
        Create a custom chart based on user configuration
        
        chart_config:
        {
            "sheet_name": "Income Statement",
            "table_id": 0,  # Which table in the sheet (optional, defaults to 0)
            "chart_type": "bar",  # bar, line, pie, scatter, area, etc.
            "x_column": "Metric",
            "y_columns": ["2023", "2024"],  # Can be multiple for comparison
            "title": "Revenue Comparison",
            "filter": {"column": "Metric", "contains": "Revenue"}  # Optional
        }
        """
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb[chart_config['sheet_name']]
        
        # Detect tables in the sheet
        tables = self._detect_tables_in_sheet(sheet)
        
        if not tables:
            raise ValueError(f"No data tables found in sheet '{chart_config['sheet_name']}'")
        
        # Get the specified table (default to first table)
        table_id = chart_config.get('table_id', 0)
        if table_id >= len(tables):
            raise ValueError(f"Table {table_id} not found. Sheet has {len(tables)} table(s)")
        
        table = tables[table_id]
        headers = table['headers']
        data = table['rows']
        
        # Apply filter if specified
        if 'filter' in chart_config and chart_config['filter']:
            filter_col = chart_config['filter']['column']
            filter_value = chart_config['filter'].get('contains', '')
            data = [row for row in data if filter_col in row and filter_value.lower() in str(row[filter_col]).lower()]
        
        # Create chart based on type
        chart_type = chart_config.get('chart_type', 'bar')
        x_column = chart_config.get('x_column')
        y_columns = chart_config.get('y_columns', [])
        title = chart_config.get('title', 'Custom Chart')
        
        # Extract x and y data
        x_data = [row.get(x_column, '') for row in data if x_column in row]
        
        if chart_type == 'bar':
            fig = go.Figure()
            for y_col in y_columns:
                y_data = [row.get(y_col, 0) for row in data]
                fig.add_trace(go.Bar(
                    name=y_col,
                    x=x_data,
                    y=y_data
                ))
            fig.update_layout(
                title=title,
                xaxis_title=x_column,
                yaxis_title='Value',
                barmode='group'
            )
        
        elif chart_type == 'line':
            fig = go.Figure()
            for y_col in y_columns:
                y_data = [row.get(y_col, 0) for row in data]
                fig.add_trace(go.Scatter(
                    name=y_col,
                    x=x_data,
                    y=y_data,
                    mode='lines+markers'
                ))
            fig.update_layout(
                title=title,
                xaxis_title=x_column,
                yaxis_title='Value'
            )
        
        elif chart_type == 'pie':
            # Use first y_column for pie chart
            y_col = y_columns[0] if y_columns else headers[1]
            y_data = [row.get(y_col, 0) for row in data]
            fig = go.Figure(data=[go.Pie(
                labels=x_data,
                values=y_data,
                hole=0.3
            )])
            fig.update_layout(title=title)
        
        elif chart_type == 'area':
            fig = go.Figure()
            for y_col in y_columns:
                y_data = [row.get(y_col, 0) for row in data]
                fig.add_trace(go.Scatter(
                    name=y_col,
                    x=x_data,
                    y=y_data,
                    fill='tonexty',
                    mode='lines'
                ))
            fig.update_layout(
                title=title,
                xaxis_title=x_column,
                yaxis_title='Value'
            )
        
        elif chart_type == 'scatter':
            # For scatter, use first two y_columns as x and y
            if len(y_columns) >= 2:
                x_scatter = [row.get(y_columns[0], 0) for row in data]
                y_scatter = [row.get(y_columns[1], 0) for row in data]
                fig = go.Figure(data=[go.Scatter(
                    x=x_scatter,
                    y=y_scatter,
                    mode='markers',
                    text=x_data,
                    marker=dict(size=10)
                )])
                fig.update_layout(
                    title=title,
                    xaxis_title=y_columns[0],
                    yaxis_title=y_columns[1]
                )
            else:
                # Fallback to bar chart
                return self.create_custom_chart(excel_path, report_id, {**chart_config, 'chart_type': 'bar'})
        
        else:
            # Default to bar chart
            return self.create_custom_chart(excel_path, report_id, {**chart_config, 'chart_type': 'bar'})
        
        # Style the chart
        fig.update_layout(
            template='plotly_white',
            height=500,
            hovermode='x unified',
            font=dict(size=12)
        )
        
        # Save chart
        filename = f"custom_{report_id}_{chart_config.get('chart_type', 'chart')}_{len(os.listdir(self.output_dir))}.html"
        output_path = os.path.join(self.output_dir, filename)
        fig.write_html(output_path)
        
        return output_path


custom_viz_service = CustomVisualizationService()
