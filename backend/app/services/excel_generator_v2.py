"""
Advanced Excel Report Generator V2
Handles nested JSON structure with multiple sheets and professional formatting
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List, Optional
import os


class ExcelGeneratorV2:
    """Generate comprehensive formatted Excel reports"""
    
    def __init__(self):
        self.output_dir = "outputs/excel"
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Define consistent styles
        self.styles = {
            'title': {
                'font': Font(bold=True, size=14, color="FFFFFF"),
                'fill': PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'header': {
                'font': Font(bold=True, size=11, color="FFFFFF"),
                'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
            },
            'subheader': {
                'font': Font(bold=True, size=10),
                'fill': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
                'alignment': Alignment(horizontal='left', vertical='center')
            },
            'label': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='left', vertical='center', indent=1)
            },
            'value': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'number_format': '#,##0.00'
            },
            'currency': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'number_format': '$#,##0'
            },
            'percentage': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'number_format': '0.00%'
            }
        }
        
        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    
    def generate_excel(self, financial_data: Dict[str, Any], output_path: str) -> str:
        """Generate comprehensive Excel report"""
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Extract data
        metadata = financial_data.get('metadata', {})
        financial_statements = financial_data.get('financial_statements', {})
        ratios = financial_data.get('financial_ratios', {})
        segments = financial_data.get('segment_analysis', [])
        geographic = financial_data.get('geographic_analysis', [])
        mgmt = financial_data.get('management_analysis', {})
        operational = financial_data.get('operational_metrics', {})
        shareholder = financial_data.get('shareholder_returns', {})
        
        # Create sheets
        print("📊 Creating Excel sheets...")
        
        self._create_summary_sheet(wb, metadata, financial_statements, ratios)
        self._create_income_statement_sheet(wb, metadata, financial_statements.get('income_statement', {}))
        self._create_balance_sheet_sheet(wb, metadata, financial_statements.get('balance_sheet', {}))
        self._create_cash_flow_sheet(wb, metadata, financial_statements.get('cash_flow', {}))
        self._create_ratios_sheet(wb, metadata, ratios)
        self._create_segments_sheet(wb, metadata, segments)
        self._create_geographic_sheet(wb, metadata, geographic)
        self._create_analysis_sheet(wb, metadata, mgmt, operational, shareholder)
        
        # Save workbook
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"✅ Excel file created: {output_path}")
        
        return output_path
    
    def _apply_style(self, cell, style_name: str):
        """Apply predefined style to cell"""
        style = self.styles.get(style_name, {})
        if 'font' in style:
            cell.font = style['font']
        if 'fill' in style:
            cell.fill = style['fill']
        if 'alignment' in style:
            cell.alignment = style['alignment']
        if 'number_format' in style:
            cell.number_format = style['number_format']
        cell.border = self.border
    
    def _write_header(self, ws, row: int, col: int, text: str, width: int = 20):
        """Write column header"""
        cell = ws.cell(row=row, column=col, value=text)
        self._apply_style(cell, 'header')
        ws.column_dimensions[get_column_letter(col)].width = width
        return cell
    
    def _write_title(self, ws, row: int, text: str, merge_cols: int = 4):
        """Write section title"""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_cols)
        cell = ws.cell(row=row, column=1, value=text)
        self._apply_style(cell, 'title')
        ws.row_dimensions[row].height = 25
        return row + 1
    
    def _write_subheader(self, ws, row: int, text: str, merge_cols: int = 4):
        """Write subsection header"""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_cols)
        cell = ws.cell(row=row, column=1, value=text)
        self._apply_style(cell, 'subheader')
        return row + 1
    
    def _create_summary_sheet(self, wb: Workbook, metadata: Dict, statements: Dict, ratios: Dict):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary")
        
        row = 1
        
        # Title
        row = self._write_title(ws, row, f"{metadata.get('company_name', 'Company')} - Financial Summary", merge_cols=4)
        row += 1
        
        # Company Info
        row = self._write_subheader(ws, row, "Company Information", merge_cols=4)
        
        info_data = [
            ("Company Name", metadata.get('company_name', 'N/A')),
            ("Fiscal Year", metadata.get('fiscal_year', 'N/A')),
            ("Reporting Period", f"{metadata.get('reporting_period_start', 'N/A')} to {metadata.get('reporting_period_end', 'N/A')}"),
            ("Currency", metadata.get('currency', 'USD')),
            ("Auditor", metadata.get('auditor_name', 'N/A'))
        ]
        
        for label, value in info_data:
            cell_label = ws.cell(row=row, column=1, value=label)
            self._apply_style(cell_label, 'label')
            cell_value = ws.cell(row=row, column=2, value=value)
            self._apply_style(cell_value, 'value')
            row += 1
        
        row += 1
        
        # Key Financial Metrics
        row = self._write_subheader(ws, row, "Key Financial Metrics (in millions)", merge_cols=4)
        
        income_current = statements.get('income_statement', {}).get('current_year', {})
        income_previous = statements.get('income_statement', {}).get('previous_year', {})
        balance_current = statements.get('balance_sheet', {}).get('current_year', {})
        cash_flow_current = statements.get('cash_flow', {}).get('current_year', {})
        
        # Headers
        self._write_header(ws, row, 1, "Metric", 30)
        self._write_header(ws, row, 2, "Current Year", 18)
        self._write_header(ws, row, 3, "Previous Year", 18)
        self._write_header(ws, row, 4, "Change %", 15)
        row += 1
        
        # Key metrics
        metrics_data = [
            ("Revenue", income_current.get('revenue'), income_previous.get('revenue')),
            ("Gross Profit", income_current.get('gross_profit'), income_previous.get('gross_profit')),
            ("Operating Income", income_current.get('operating_income'), income_previous.get('operating_income')),
            ("Net Income", income_current.get('net_income'), income_previous.get('net_income')),
            ("EPS (Diluted)", income_current.get('diluted_eps'), income_previous.get('diluted_eps')),
            ("Total Assets", balance_current.get('total_assets'), statements.get('balance_sheet', {}).get('previous_year', {}).get('total_assets')),
            ("Total Liabilities", balance_current.get('total_liabilities'), statements.get('balance_sheet', {}).get('previous_year', {}).get('total_liabilities')),
            ("Shareholders Equity", balance_current.get('total_shareholders_equity'), statements.get('balance_sheet', {}).get('previous_year', {}).get('total_shareholders_equity')),
            ("Operating Cash Flow", cash_flow_current.get('net_cash_from_operating_activities'), statements.get('cash_flow', {}).get('previous_year', {}).get('net_cash_from_operating_activities')),
            ("Free Cash Flow", cash_flow_current.get('free_cash_flow'), statements.get('cash_flow', {}).get('previous_year', {}).get('free_cash_flow'))
        ]
        
        for label, current, previous in metrics_data:
            cell_label = ws.cell(row=row, column=1, value=label)
            self._apply_style(cell_label, 'label')
            
            cell_current = ws.cell(row=row, column=2, value=current if current else None)
            self._apply_style(cell_current, 'currency')
            
            cell_previous = ws.cell(row=row, column=3, value=previous if previous else None)
            self._apply_style(cell_previous, 'currency')
            
            # Calculate change %
            if current and previous and previous != 0:
                change_pct = ((current - previous) / previous)
                cell_change = ws.cell(row=row, column=4, value=change_pct)
                self._apply_style(cell_change, 'percentage')
                # Color code: green for positive, red for negative
                if change_pct > 0:
                    cell_change.font = Font(color="008000", size=10)
                elif change_pct < 0:
                    cell_change.font = Font(color="FF0000", size=10)
            
            row += 1
        
        row += 1
        
        # Key Ratios
        row = self._write_subheader(ws, row, "Key Financial Ratios", merge_cols=4)
        
        # Extract ratios (handle nested structure)
        gross_margin = ratios.get('gross_margin')
        if gross_margin is None:
            gross_margin = ratios.get('profitability', {}).get('gross_margin')
        
        operating_margin = ratios.get('operating_margin')
        if operating_margin is None:
            operating_margin = ratios.get('profitability', {}).get('operating_margin')
        
        net_margin = ratios.get('net_profit_margin')
        if net_margin is None:
            net_margin = ratios.get('profitability', {}).get('net_profit_margin')
        
        roe = ratios.get('roe')
        if roe is None:
            roe = ratios.get('profitability', {}).get('roe')
        
        roa = ratios.get('roa')
        if roa is None:
            roa = ratios.get('profitability', {}).get('roa')
        
        current_ratio = ratios.get('current_ratio')
        if current_ratio is None:
            current_ratio = ratios.get('liquidity', {}).get('current_ratio')
        
        debt_to_equity = ratios.get('debt_to_equity')
        if debt_to_equity is None:
            debt_to_equity = ratios.get('leverage', {}).get('debt_to_equity')
        
        ratios_data = [
            ("Gross Margin", gross_margin, True),
            ("Operating Margin", operating_margin, True),
            ("Net Profit Margin", net_margin, True),
            ("Return on Equity (ROE)", roe, True),
            ("Return on Assets (ROA)", roa, True),
            ("Current Ratio", current_ratio, False),
            ("Debt to Equity", debt_to_equity, False)
        ]
        
        for label, value, is_percentage in ratios_data:
            cell_label = ws.cell(row=row, column=1, value=label)
            self._apply_style(cell_label, 'label')
            
            if value is not None:
                # Convert to percentage if needed
                if is_percentage and value < 10:  # Likely a decimal (0.3596)
                    value = value / 100
                elif is_percentage and value >= 10:  # Already a percentage (35.96)
                    value = value / 100
                
                cell_value = ws.cell(row=row, column=2, value=value)
                self._apply_style(cell_value, 'percentage' if is_percentage else 'value')
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
    
    def _create_income_statement_sheet(self, wb: Workbook, metadata: Dict, income_stmt: Dict):
        """Create detailed income statement"""
        ws = wb.create_sheet("Income Statement")
        
        row = 1
        row = self._write_title(ws, row, f"{metadata.get('company_name', 'Company')} - Income Statement", merge_cols=3)
        row += 1
        
        # Headers
        self._write_header(ws, row, 1, "Line Item", 35)
        self._write_header(ws, row, 2, f"FY {metadata.get('fiscal_year', 'Current')}", 18)
        self._write_header(ws, row, 3, f"FY {metadata.get('fiscal_year', 2023) - 1}", 18)
        row += 1
        
        current = income_stmt.get('current_year', {})
        previous = income_stmt.get('previous_year', {})
        
        # Income statement line items
        line_items = [
            ("Revenue", 'revenue', False),
            ("Cost of Revenue", 'cost_of_revenue', True),
            ("Gross Profit", 'gross_profit', False),
            ("", None, False),  # Blank row
            ("Operating Expenses:", None, False),
            ("  Research & Development", 'research_and_development', True),
            ("  Sales & Marketing", 'sales_and_marketing', True),
            ("  General & Administrative", 'general_and_administrative', True),
            ("Total Operating Expenses", 'total_operating_expenses', True),
            ("Operating Income", 'operating_income', False),
            ("", None, False),
            ("Other Income/Expense:", None, False),
            ("  Interest Income", 'interest_income', False),
            ("  Interest Expense", 'interest_expense', True),
            ("  Other Income (Expense)", 'other_income_expense', False),
            ("Income Before Tax", 'income_before_tax', False),
            ("  Income Tax Expense", 'income_tax_expense', True),
            ("Net Income", 'net_income', False),
            ("", None, False),
            ("Per Share Data:", None, False),
            ("  Basic EPS", 'basic_eps', False),
            ("  Diluted EPS", 'diluted_eps', False),
            ("  Shares Outstanding (Basic)", 'weighted_average_shares_basic', False),
            ("  Shares Outstanding (Diluted)", 'weighted_average_shares_diluted', False),
        ]
        
        for label, key, is_negative in line_items:
            if key is None:
                # Section header or blank row
                if label:
                    cell = ws.cell(row=row, column=1, value=label)
                    self._apply_style(cell, 'subheader')
                row += 1
                continue
            
            cell_label = ws.cell(row=row, column=1, value=label)
            self._apply_style(cell_label, 'label')
            
            current_val = current.get(key)
            previous_val = previous.get(key)
            
            # Apply negative sign if needed
            if is_negative:
                if current_val:
                    current_val = -abs(current_val)
                if previous_val:
                    previous_val = -abs(previous_val)
            
            cell_current = ws.cell(row=row, column=2, value=current_val)
            cell_previous = ws.cell(row=row, column=3, value=previous_val)
            
            # Determine number format (currency for most, number for shares/EPS)
            if 'shares' in key.lower() or 'eps' in key.lower():
                self._apply_style(cell_current, 'value')
                self._apply_style(cell_previous, 'value')
            else:
                self._apply_style(cell_current, 'currency')
                self._apply_style(cell_previous, 'currency')
            
            row += 1
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
    
    def _create_balance_sheet_sheet(self, wb: Workbook, metadata: Dict, balance_sheet: Dict):
        """Create detailed balance sheet"""
        ws = wb.create_sheet("Balance Sheet")
        
        row = 1
        row = self._write_title(ws, row, f"{metadata.get('company_name', 'Company')} - Balance Sheet", merge_cols=3)
        row += 1
        
        # Headers
        self._write_header(ws, row, 1, "Line Item", 40)
        self._write_header(ws, row, 2, f"FY {metadata.get('fiscal_year', 'Current')}", 18)
        self._write_header(ws, row, 3, f"FY {metadata.get('fiscal_year', 2023) - 1}", 18)
        row += 1
        
        current = balance_sheet.get('current_year', {})
        previous = balance_sheet.get('previous_year', {})
        
        # Balance sheet line items
        line_items = [
            ("ASSETS", None, False, True),
            ("Current Assets:", None, False, False),
            ("  Cash and Cash Equivalents", 'cash_and_cash_equivalents', False, False),
            ("  Short-term Investments", 'short_term_investments', False, False),
            ("  Accounts Receivable", 'accounts_receivable', False, False),
            ("  Inventory", 'inventory', False, False),
            ("  Other Current Assets", 'other_current_assets', False, False),
            ("Total Current Assets", 'total_current_assets', False, True),
            ("", None, False, False),
            ("Non-Current Assets:", None, False, False),
            ("  Property, Plant & Equipment (Net)", 'property_plant_equipment_net', False, False),
            ("  Intangible Assets", 'intangible_assets', False, False),
            ("  Goodwill", 'goodwill', False, False),
            ("  Long-term Investments", 'long_term_investments', False, False),
            ("  Other Non-Current Assets", 'other_non_current_assets', False, False),
            ("Total Non-Current Assets", 'total_non_current_assets', False, True),
            ("TOTAL ASSETS", 'total_assets', False, True),
            ("", None, False, False),
            ("LIABILITIES", None, False, True),
            ("Current Liabilities:", None, False, False),
            ("  Accounts Payable", 'accounts_payable', False, False),
            ("  Short-term Debt", 'short_term_debt', False, False),
            ("  Deferred Revenue (Current)", 'deferred_revenue_current', False, False),
            ("  Current Portion of Long-term Debt", 'current_portion_long_term_debt', False, False),
            ("  Other Current Liabilities", 'other_current_liabilities', False, False),
            ("Total Current Liabilities", 'total_current_liabilities', False, True),
            ("", None, False, False),
            ("Non-Current Liabilities:", None, False, False),
            ("  Long-term Debt", 'long_term_debt', False, False),
            ("  Deferred Tax Liabilities", 'deferred_tax_liabilities', False, False),
            ("  Deferred Revenue (Non-Current)", 'deferred_revenue_non_current', False, False),
            ("  Other Long-term Liabilities", 'other_long_term_liabilities', False, False),
            ("Total Non-Current Liabilities", 'total_non_current_liabilities', False, True),
            ("TOTAL LIABILITIES", 'total_liabilities', False, True),
            ("", None, False, False),
            ("SHAREHOLDERS' EQUITY", None, False, True),
            ("  Additional Paid-in Capital", 'additional_paid_in_capital', False, False),
            ("  Retained Earnings", 'retained_earnings', False, False),
            ("  Accumulated Other Comprehensive Income", 'accumulated_other_comprehensive_income', False, False),
            ("TOTAL SHAREHOLDERS' EQUITY", 'total_shareholders_equity', False, True),
            ("", None, False, False),
            ("TOTAL LIABILITIES & EQUITY", 'total_liabilities', False, True),  # Will calculate
        ]
        
        for label, key, is_negative, is_bold in line_items:
            if key is None:
                if label:
                    cell = ws.cell(row=row, column=1, value=label)
                    if is_bold:
                        self._apply_style(cell, 'subheader')
                    else:
                        self._apply_style(cell, 'label')
                        cell.font = Font(bold=True, size=10)
                row += 1
                continue
            
            cell_label = ws.cell(row=row, column=1, value=label)
            self._apply_style(cell_label, 'label')
            if is_bold:
                cell_label.font = Font(bold=True, size=10)
            
            current_val = current.get(key)
            previous_val = previous.get(key)
            
            # Special case: Total Liabilities & Equity
            if label == "TOTAL LIABILITIES & EQUITY":
                current_val = current.get('total_liabilities', 0) + current.get('total_shareholders_equity', 0) if current.get('total_liabilities') and current.get('total_shareholders_equity') else None
                previous_val = previous.get('total_liabilities', 0) + previous.get('total_shareholders_equity', 0) if previous.get('total_liabilities') and previous.get('total_shareholders_equity') else None
            
            cell_current = ws.cell(row=row, column=2, value=current_val)
            self._apply_style(cell_current, 'currency')
            if is_bold:
                cell_current.font = Font(bold=True, size=10)
            
            cell_previous = ws.cell(row=row, column=3, value=previous_val)
            self._apply_style(cell_previous, 'currency')
            if is_bold:
                cell_previous.font = Font(bold=True, size=10)
            
            row += 1
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
    
    def _create_cash_flow_sheet(self, wb: Workbook, metadata: Dict, cash_flow: Dict):
        """Create cash flow statement"""
        ws = wb.create_sheet("Cash Flow")
        
        row = 1
        row = self._write_title(ws, row, f"{metadata.get('company_name', 'Company')} - Cash Flow Statement", merge_cols=3)
        row += 1
        
        # Headers
        self._write_header(ws, row, 1, "Line Item", 45)
        self._write_header(ws, row, 2, f"FY {metadata.get('fiscal_year', 'Current')}", 18)
        self._write_header(ws, row, 3, f"FY {metadata.get('fiscal_year', 2023) - 1}", 18)
        row += 1
        
        current = cash_flow.get('current_year', {})
        previous = cash_flow.get('previous_year', {})
        
        line_items = [
            ("OPERATING ACTIVITIES", None, False, True),
            ("  Net Income", 'net_income_cf', False, False),
            ("  Depreciation & Amortization", 'depreciation_amortization', False, False),
            ("  Stock-based Compensation", 'stock_based_compensation', False, False),
            ("  Deferred Income Taxes", 'deferred_income_taxes', False, False),
            ("  Changes in Working Capital", 'changes_in_working_capital', False, False),
            ("  Other Operating Activities", 'other_operating_activities', False, False),
            ("Net Cash from Operating Activities", 'net_cash_from_operating_activities', False, True),
            ("", None, False, False),
            ("INVESTING ACTIVITIES", None, False, True),
            ("  Capital Expenditures", 'capital_expenditures', True, False),
            ("  Acquisitions", 'acquisitions', True, False),
            ("  Purchases of Investments", 'purchases_of_investments', True, False),
            ("  Sales of Investments", 'sales_of_investments', False, False),
            ("  Other Investing Activities", 'other_investing_activities', False, False),
            ("Net Cash from Investing Activities", 'net_cash_from_investing_activities', False, True),
            ("", None, False, False),
            ("FINANCING ACTIVITIES", None, False, True),
            ("  Proceeds from Debt", 'proceeds_from_debt', False, False),
            ("  Repayment of Debt", 'repayment_of_debt', True, False),
            ("  Dividends Paid", 'dividends_paid', True, False),
            ("  Stock Repurchases", 'stock_repurchases', True, False),
            ("  Proceeds from Stock Issuance", 'proceeds_from_stock_issuance', False, False),
            ("  Other Financing Activities", 'other_financing_activities', False, False),
            ("Net Cash from Financing Activities", 'net_cash_from_financing_activities', False, True),
            ("", None, False, False),
            ("Net Change in Cash", 'net_change_in_cash', False, True),
            ("Cash - Beginning of Period", 'cash_beginning_of_period', False, False),
            ("Cash - End of Period", 'cash_end_of_period', False, True),
            ("", None, False, False),
            ("FREE CASH FLOW", 'free_cash_flow', False, True),
        ]
        
        for label, key, is_negative, is_bold in line_items:
            if key is None:
                if label:
                    cell = ws.cell(row=row, column=1, value=label)
                    if is_bold:
                        self._apply_style(cell, 'subheader')
                    else:
                        self._apply_style(cell, 'label')
                row += 1
                continue
            
            cell_label = ws.cell(row=row, column=1, value=label)
            self._apply_style(cell_label, 'label')
            if is_bold:
                cell_label.font = Font(bold=True, size=10)
            
            current_val = current.get(key)
            previous_val = previous.get(key)
            
            # Some values should be shown as negative
            if is_negative:
                if current_val and current_val > 0:
                    current_val = -current_val
                if previous_val and previous_val > 0:
                    previous_val = -previous_val
            
            cell_current = ws.cell(row=row, column=2, value=current_val)
            self._apply_style(cell_current, 'currency')
            if is_bold:
                cell_current.font = Font(bold=True, size=10)
            
            cell_previous = ws.cell(row=row, column=3, value=previous_val)
            self._apply_style(cell_previous, 'currency')
            if is_bold:
                cell_previous.font = Font(bold=True, size=10)
            
            row += 1
        
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
    
    def _create_ratios_sheet(self, wb: Workbook, metadata: Dict, ratios: Dict):
        """Create financial ratios analysis sheet"""
        ws = wb.create_sheet("Financial Ratios")
        
        row = 1
        row = self._write_title(ws, row, f"{metadata.get('company_name', 'Company')} - Financial Ratios", merge_cols=2)
        row += 1
        
        self._write_header(ws, row, 1, "Ratio Category / Metric", 40)
        self._write_header(ws, row, 2, "Value", 20)
        row += 1
        
        # Handle both nested and flat structures
        profitability = ratios.get('profitability', {})
        liquidity = ratios.get('liquidity', {})
        leverage = ratios.get('leverage', {})
        efficiency = ratios.get('efficiency', {})
        growth = ratios.get('growth', {})
        per_share = ratios.get('per_share', {})
        
        # Fallback to flat structure if nested is empty
        if not profitability:
            profitability = {
                'gross_margin': ratios.get('gross_margin'),
                'operating_margin': ratios.get('operating_margin'),
                'net_profit_margin': ratios.get('net_profit_margin'),
                'roe': ratios.get('roe'),
                'roa': ratios.get('roa'),
            }
        
        if not liquidity:
            liquidity = {
                'current_ratio': ratios.get('current_ratio'),
                'working_capital': ratios.get('working_capital'),
            }
        
        if not leverage:
            leverage = {
                'debt_to_equity': ratios.get('debt_to_equity'),
                'debt_to_assets': ratios.get('debt_to_assets'),
            }
        
        ratio_sections = [
            ("PROFITABILITY RATIOS", [
                ("Gross Margin", profitability.get('gross_margin'), True),
                ("Operating Margin", profitability.get('operating_margin'), True),
                ("Net Profit Margin", profitability.get('net_profit_margin'), True),
                ("Return on Equity (ROE)", profitability.get('roe'), True),
                ("Return on Assets (ROA)", profitability.get('roa'), True),
                ("Return on Invested Capital (ROIC)", profitability.get('roic'), True),
            ]),
            ("LIQUIDITY RATIOS", [
                ("Current Ratio", liquidity.get('current_ratio'), False),
                ("Quick Ratio", liquidity.get('quick_ratio'), False),
                ("Cash Ratio", liquidity.get('cash_ratio'), False),
                ("Working Capital (M)", liquidity.get('working_capital'), False),
            ]),
            ("LEVERAGE RATIOS", [
                ("Debt to Equity", leverage.get('debt_to_equity'), False),
                ("Debt to Assets", leverage.get('debt_to_assets'), False),
                ("Interest Coverage", leverage.get('interest_coverage'), False),
            ]),
            ("EFFICIENCY RATIOS", [
                ("Asset Turnover", efficiency.get('asset_turnover'), False),
                ("Inventory Turnover", efficiency.get('inventory_turnover'), False),
                ("Receivables Turnover", efficiency.get('receivables_turnover'), False),
            ]),
            ("GROWTH RATES", [
                ("Revenue Growth YoY", growth.get('revenue_growth_yoy'), True),
                ("Net Income Growth YoY", growth.get('net_income_growth_yoy'), True),
                ("EPS Growth YoY", growth.get('eps_growth_yoy'), True),
            ]),
            ("PER SHARE METRICS", [
                ("Book Value per Share", per_share.get('book_value_per_share'), False),
                ("Revenue per Share", per_share.get('revenue_per_share'), False),
                ("Cash per Share", per_share.get('cash_per_share'), False),
            ])
        ]
        
        for section_title, metrics in ratio_sections:
            row = self._write_subheader(ws, row, section_title, merge_cols=2)
            
            for label, value, is_percentage in metrics:
                if value is None:
                    continue
                
                cell_label = ws.cell(row=row, column=1, value=f"  {label}")
                self._apply_style(cell_label, 'label')
                
                # Convert to percentage format if needed
                if is_percentage and value < 10:  # Decimal format
                    value = value / 100
                elif is_percentage and value >= 10:  # Already percentage
                    value = value / 100
                
                cell_value = ws.cell(row=row, column=2, value=value)
                self._apply_style(cell_value, 'percentage' if is_percentage else 'value')
                
                row += 1
            
            row += 1
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
    
    def _create_segments_sheet(self, wb: Workbook, metadata: Dict, segments: List[Dict]):
        """Create business segments analysis"""
        ws = wb.create_sheet("Segment Analysis")
        
        row = 1
        row = self._write_title(ws, row, f"{metadata.get('company_name', 'Company')} - Business Segments", merge_cols=5)
        row += 1
        
        if not segments:
            ws.cell(row=row, column=1, value="No segment data available")
            return
        
        # Headers
        self._write_header(ws, row, 1, "Segment Name", 30)
        self._write_header(ws, row, 2, "Revenue (Current)", 18)
        self._write_header(ws, row, 3, "Revenue (Previous)", 18)
        self._write_header(ws, row, 4, "Operating Income", 18)
        self._write_header(ws, row, 5, "Margin %", 15)
        row += 1
        
        for segment in segments:
            cell_name = ws.cell(row=row, column=1, value=segment.get('segment_name', 'N/A'))
            self._apply_style(cell_name, 'label')
            
            cell_rev_curr = ws.cell(row=row, column=2, value=segment.get('revenue_current'))
            self._apply_style(cell_rev_curr, 'currency')
            
            cell_rev_prev = ws.cell(row=row, column=3, value=segment.get('revenue_previous'))
            self._apply_style(cell_rev_prev, 'currency')
            
            cell_op_income = ws.cell(row=row, column=4, value=segment.get('operating_income_current'))
            self._apply_style(cell_op_income, 'currency')
            
            margin = segment.get('segment_margin')
            if margin and margin < 1:  # Decimal format
                margin = margin
            elif margin and margin >= 1:  # Percentage format
                margin = margin / 100
            
            cell_margin = ws.cell(row=row, column=5, value=margin)
            self._apply_style(cell_margin, 'percentage')
            
            row += 1
        
        ws.column_dimensions['A'].width = 30
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 18
        ws.column_dimensions['E'].width = 15
    
    def _create_geographic_sheet(self, wb: Workbook, metadata: Dict, geographic: List[Dict]):
        """Create geographic analysis"""
        ws = wb.create_sheet("Geographic Analysis")
        
        row = 1
        row = self._write_title(ws, row, f"{metadata.get('company_name', 'Company')} - Geographic Breakdown", merge_cols=4)
        row += 1
        
        if not geographic:
            ws.cell(row=row, column=1, value="No geographic data available")
            return
        
        # Headers
        self._write_header(ws, row, 1, "Region", 30)
        self._write_header(ws, row, 2, "Revenue (Current)", 18)
        self._write_header(ws, row, 3, "Revenue (Previous)", 18)
        self._write_header(ws, row, 4, "YoY Growth %", 15)
        row += 1
        
        for region in geographic:
            cell_region = ws.cell(row=row, column=1, value=region.get('region_name', 'N/A'))
            self._apply_style(cell_region, 'label')
            
            rev_curr = region.get('revenue_current')
            rev_prev = region.get('revenue_previous')
            
            cell_curr = ws.cell(row=row, column=2, value=rev_curr)
            self._apply_style(cell_curr, 'currency')
            
            cell_prev = ws.cell(row=row, column=3, value=rev_prev)
            self._apply_style(cell_prev, 'currency')
            
            # Calculate growth
            if rev_curr and rev_prev and rev_prev != 0:
                growth = (rev_curr - rev_prev) / rev_prev
                cell_growth = ws.cell(row=row, column=4, value=growth)
                self._apply_style(cell_growth, 'percentage')
            
            row += 1
        
        ws.column_dimensions['A'].width = 30
        for col in ['B', 'C']:
            ws.column_dimensions[col].width = 18
        ws.column_dimensions['D'].width = 15
    
    def _create_analysis_sheet(self, wb: Workbook, metadata: Dict, mgmt: Dict, operational: Dict, shareholder: Dict):
        """Create management analysis and other qualitative data"""
        ws = wb.create_sheet("Analysis & Insights")
        
        row = 1
        row = self._write_title(ws, row, f"{metadata.get('company_name', 'Company')} - Management Analysis", merge_cols=2)
        row += 1
        
        # Management Analysis
        if mgmt:
            row = self._write_subheader(ws, row, "Business Overview", merge_cols=2)
            ws.merge_cells(start_row=row, start_column=1, end_row=row+2, end_column=2)
            cell = ws.cell(row=row, column=1, value=mgmt.get('business_overview', 'N/A'))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 60
            row += 3
            
            row = self._write_subheader(ws, row, "Key Strategies", merge_cols=2)
            ws.merge_cells(start_row=row, start_column=1, end_row=row+3, end_column=2)
            cell = ws.cell(row=row, column=1, value=mgmt.get('key_strategies', 'N/A'))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 80
            row += 4
            
            row = self._write_subheader(ws, row, "Key Risks", merge_cols=2)
            ws.merge_cells(start_row=row, start_column=1, end_row=row+2, end_column=2)
            cell = ws.cell(row=row, column=1, value=mgmt.get('key_risks', 'N/A'))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 60
            row += 3
        
        # Operational Metrics
        if operational:
            row = self._write_subheader(ws, row, "Operational Metrics", merge_cols=2)
            
            metrics = [
                ("Employee Count", operational.get('employee_count')),
                ("Employee Growth Rate", operational.get('employee_growth_rate')),
                ("Customer Count", operational.get('customer_count')),
            ]
            
            for label, value in metrics:
                if value:
                    cell_label = ws.cell(row=row, column=1, value=label)
                    self._apply_style(cell_label, 'label')
                    cell_value = ws.cell(row=row, column=2, value=value)
                    self._apply_style(cell_value, 'value')
                    row += 1
            
            row += 1
        
        # Shareholder Returns
        if shareholder:
            row = self._write_subheader(ws, row, "Shareholder Returns", merge_cols=2)
            
            metrics = [
                ("Dividend per Share", shareholder.get('dividend_per_share')),
                ("Total Dividends Paid", shareholder.get('total_dividends_paid')),
                ("Share Repurchases", shareholder.get('share_repurchases')),
            ]
            
            for label, value in metrics:
                if value:
                    cell_label = ws.cell(row=row, column=1, value=label)
                    self._apply_style(cell_label, 'label')
                    cell_value = ws.cell(row=row, column=2, value=value)
                    self._apply_style(cell_value, 'value')
                    row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50


# Create singleton instance
excel_generator_v2 = ExcelGeneratorV2()
