"""
Test script to generate visualizations from Excel file
Reads data from the generated Excel sheets and creates Plotly charts
"""
import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import json

# Path to the Excel file
excel_file = "outputs/excel/Microsoft_Corporation_FY2024_Financial_Report.xlsx"
output_dir = "outputs/visualizations"
os.makedirs(output_dir, exist_ok=True)

print(f"📊 Reading Excel file: {excel_file}")

# Load the workbook
wb = openpyxl.load_workbook(excel_file)
print(f"✅ Loaded workbook with {len(wb.sheetnames)} sheets")
print(f"   Sheets: {wb.sheetnames}")

# ===== 1. REVENUE COMPARISON (from Executive Summary) =====
print("\n📈 Creating Revenue Comparison Chart...")

summary_sheet = wb["Executive Summary"]
company_name = summary_sheet['B1'].value  # Assuming company name is here

# Find the key metrics section
revenue_current = None
revenue_previous = None
net_income_current = None
net_income_previous = None

# Scan for Revenue and Net Income rows
for row in summary_sheet.iter_rows(min_row=1, max_row=50):
    if row[0].value and 'Revenue' in str(row[0].value):
        revenue_current = row[1].value if row[1].value else 0
        revenue_previous = row[2].value if row[2].value else 0
    if row[0].value and 'Net Income' in str(row[0].value) and 'Per Share' not in str(row[0].value):
        net_income_current = row[1].value if row[1].value else 0
        net_income_previous = row[2].value if row[2].value else 0

if revenue_current and revenue_previous:
    fig = go.Figure()
    fig.add_trace(go.Bar(
        name='Revenue',
        x=['Previous Year', 'Current Year'],
        y=[revenue_previous, revenue_current],
        marker_color='rgb(26, 118, 255)'
    ))
    if net_income_current and net_income_previous:
        fig.add_trace(go.Bar(
            name='Net Income',
            x=['Previous Year', 'Current Year'],
            y=[net_income_previous, net_income_current],
            marker_color='rgb(55, 83, 109)'
        ))
    
    fig.update_layout(
        title='Revenue & Net Income Comparison',
        xaxis_title='Period',
        yaxis_title='Amount (Millions)',
        barmode='group',
        template='plotly_white'
    )
    
    output_file = f"{output_dir}/revenue_comparison_from_excel.html"
    fig.write_html(output_file)
    print(f"✅ Created: {output_file}")

# ===== 2. FINANCIAL RATIOS DASHBOARD (from Financial Ratios sheet) =====
print("\n📊 Creating Financial Ratios Dashboard...")

ratios_sheet = wb["Financial Ratios"]
ratios_data = {}

# Scan ratios sheet for data
for row in ratios_sheet.iter_rows(min_row=1, max_row=100):
    if row[0].value and row[1].value:
        label = str(row[0].value).strip()
        try:
            value = float(row[1].value)
            # Skip section headers
            if not label.isupper() and not label.endswith(':'):
                ratios_data[label] = value
        except (ValueError, TypeError):
            continue

if ratios_data:
    # Create subplots for different ratio categories
    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=('Profitability', 'Liquidity', 'Leverage', 'Growth'),
        specs=[[{'type': 'indicator'}, {'type': 'indicator'}],
               [{'type': 'indicator'}, {'type': 'indicator'}]]
    )
    
    # Profitability - ROE
    roe = ratios_data.get('Return on Equity (ROE)', 0)
    if roe > 1:  # If it's a percentage value
        roe = roe / 100
    fig.add_trace(go.Indicator(
        mode="gauge+number+delta",
        value=roe * 100,
        title={'text': "ROE (%)"},
        gauge={'axis': {'range': [0, 100]},
               'bar': {'color': "darkblue"},
               'steps': [
                   {'range': [0, 10], 'color': "lightgray"},
                   {'range': [10, 20], 'color': "gray"}],
               'threshold': {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': 15}}
    ), row=1, col=1)
    
    # Liquidity - Current Ratio
    current_ratio = ratios_data.get('Current Ratio', 0)
    fig.add_trace(go.Indicator(
        mode="gauge+number",
        value=current_ratio,
        title={'text': "Current Ratio"},
        gauge={'axis': {'range': [0, 3]},
               'bar': {'color': "green"},
               'steps': [
                   {'range': [0, 1], 'color': "lightcoral"},
                   {'range': [1, 2], 'color': "lightyellow"}],
               'threshold': {'line': {'color': "darkgreen", 'width': 4}, 'thickness': 0.75, 'value': 1.5}}
    ), row=1, col=2)
    
    # Leverage - Debt to Equity
    debt_to_equity = ratios_data.get('Debt to Equity', 0)
    fig.add_trace(go.Indicator(
        mode="gauge+number",
        value=debt_to_equity,
        title={'text': "Debt/Equity"},
        gauge={'axis': {'range': [0, 2]},
               'bar': {'color': "orange"},
               'steps': [
                   {'range': [0, 0.5], 'color': "lightgreen"},
                   {'range': [0.5, 1], 'color': "lightyellow"}],
               'threshold': {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': 1}}
    ), row=2, col=1)
    
    # Growth - Revenue Growth
    revenue_growth = ratios_data.get('Revenue Growth YoY', 0)
    if revenue_growth < 1 and revenue_growth != 0:  # Decimal format
        revenue_growth = revenue_growth * 100
    fig.add_trace(go.Indicator(
        mode="number+delta",
        value=revenue_growth,
        title={'text': "Revenue Growth (%)"},
        number={'suffix': "%"},
        delta={'reference': 10, 'relative': False}
    ), row=2, col=2)
    
    fig.update_layout(
        height=600,
        title_text="Key Financial Ratios Dashboard",
        template='plotly_white'
    )
    
    output_file = f"{output_dir}/ratios_dashboard_from_excel.html"
    fig.write_html(output_file)
    print(f"✅ Created: {output_file}")

# ===== 3. BALANCE SHEET COMPOSITION (from Balance Sheet) =====
print("\n📊 Creating Balance Sheet Visualization...")

balance_sheet = wb["Balance Sheet"]

# Find key balance sheet items
total_current_assets = None
total_non_current_assets = None
total_current_liabilities = None
total_non_current_liabilities = None
total_equity = None

for row in balance_sheet.iter_rows(min_row=1, max_row=100):
    if row[0].value:
        label = str(row[0].value).strip()
        try:
            if label == "Total Current Assets":
                total_current_assets = float(row[1].value) if row[1].value else 0
            elif label == "Total Non-Current Assets":
                total_non_current_assets = float(row[1].value) if row[1].value else 0
            elif label == "Total Current Liabilities":
                total_current_liabilities = float(row[1].value) if row[1].value else 0
            elif label == "Total Non-Current Liabilities":
                total_non_current_liabilities = float(row[1].value) if row[1].value else 0
            elif label == "TOTAL SHAREHOLDERS' EQUITY":
                total_equity = float(row[1].value) if row[1].value else 0
        except (ValueError, TypeError):
            continue

if total_current_assets and total_non_current_assets:
    # Create side-by-side bar chart for Assets vs Liabilities & Equity
    fig = go.Figure()
    
    fig.add_trace(go.Bar(
        name='Current Assets',
        x=['Assets'],
        y=[total_current_assets],
        marker_color='lightblue'
    ))
    fig.add_trace(go.Bar(
        name='Non-Current Assets',
        x=['Assets'],
        y=[total_non_current_assets],
        marker_color='darkblue'
    ))
    fig.add_trace(go.Bar(
        name='Current Liabilities',
        x=['Liabilities & Equity'],
        y=[total_current_liabilities],
        marker_color='lightcoral'
    ))
    fig.add_trace(go.Bar(
        name='Non-Current Liabilities',
        x=['Liabilities & Equity'],
        y=[total_non_current_liabilities],
        marker_color='darkred'
    ))
    fig.add_trace(go.Bar(
        name='Shareholders Equity',
        x=['Liabilities & Equity'],
        y=[total_equity],
        marker_color='lightgreen'
    ))
    
    fig.update_layout(
        title='Balance Sheet Composition',
        yaxis_title='Amount (Millions)',
        barmode='stack',
        template='plotly_white',
        height=600
    )
    
    output_file = f"{output_dir}/balance_sheet_from_excel.html"
    fig.write_html(output_file)
    print(f"✅ Created: {output_file}")

# ===== 4. CASH FLOW WATERFALL (from Cash Flow sheet) =====
print("\n💰 Creating Cash Flow Waterfall...")

cash_flow_sheet = wb["Cash Flow"]

operating_cf = None
investing_cf = None
financing_cf = None
net_change = None

for row in cash_flow_sheet.iter_rows(min_row=1, max_row=100):
    if row[0].value:
        label = str(row[0].value).strip()
        try:
            if label == "Net Cash from Operating Activities":
                operating_cf = float(row[1].value) if row[1].value else 0
            elif label == "Net Cash from Investing Activities":
                investing_cf = float(row[1].value) if row[1].value else 0
            elif label == "Net Cash from Financing Activities":
                financing_cf = float(row[1].value) if row[1].value else 0
            elif label == "Net Change in Cash":
                net_change = float(row[1].value) if row[1].value else 0
        except (ValueError, TypeError):
            continue

if operating_cf and investing_cf and financing_cf:
    fig = go.Figure(go.Waterfall(
        name="Cash Flow",
        orientation="v",
        measure=["relative", "relative", "relative", "total"],
        x=["Operating Activities", "Investing Activities", "Financing Activities", "Net Change"],
        textposition="outside",
        text=[f"${operating_cf:,.0f}M", f"${investing_cf:,.0f}M", f"${financing_cf:,.0f}M", f"${net_change:,.0f}M"],
        y=[operating_cf, investing_cf, financing_cf, net_change],
        connector={"line": {"color": "rgb(63, 63, 63)"}},
        decreasing={"marker": {"color": "red"}},
        increasing={"marker": {"color": "green"}},
        totals={"marker": {"color": "blue"}}
    ))
    
    fig.update_layout(
        title="Cash Flow Waterfall Chart",
        yaxis_title="Amount (Millions)",
        template='plotly_white',
        height=600
    )
    
    output_file = f"{output_dir}/cash_flow_waterfall_from_excel.html"
    fig.write_html(output_file)
    print(f"✅ Created: {output_file}")

# ===== 5. SEGMENT ANALYSIS (from Segment Analysis sheet) =====
print("\n🔍 Creating Segment Analysis...")

if "Segment Analysis" in wb.sheetnames:
    segment_sheet = wb["Segment Analysis"]
    
    segments = []
    revenues = []
    margins = []
    
    # Skip header row, read data
    for row in segment_sheet.iter_rows(min_row=4, max_row=20):  # Assuming data starts at row 4
        if row[0].value and row[1].value:
            try:
                segment_name = str(row[0].value).strip()
                revenue = float(row[1].value) if row[1].value else 0
                margin = float(row[4].value) if row[4].value else 0
                
                if revenue > 0:  # Only add if there's actual data
                    segments.append(segment_name)
                    revenues.append(revenue)
                    # Convert margin to percentage if needed
                    if margin < 1:
                        margin = margin * 100
                    margins.append(margin)
            except (ValueError, TypeError):
                continue
    
    if segments and revenues:
        # Create subplot with bar chart and line chart
        fig = make_subplots(specs=[[{"secondary_y": True}]])
        
        fig.add_trace(
            go.Bar(name="Revenue", x=segments, y=revenues, marker_color='skyblue'),
            secondary_y=False
        )
        
        fig.add_trace(
            go.Scatter(name="Margin %", x=segments, y=margins, mode='lines+markers', 
                      marker=dict(size=10, color='red'), line=dict(color='red', width=3)),
            secondary_y=True
        )
        
        fig.update_xaxes(title_text="Business Segment")
        fig.update_yaxes(title_text="Revenue (Millions)", secondary_y=False)
        fig.update_yaxes(title_text="Margin (%)", secondary_y=True)
        
        fig.update_layout(
            title="Business Segment Performance",
            template='plotly_white',
            height=600
        )
        
        output_file = f"{output_dir}/segment_analysis_from_excel.html"
        fig.write_html(output_file)
        print(f"✅ Created: {output_file}")

# ===== 6. GEOGRAPHIC BREAKDOWN (from Geographic Analysis sheet) =====
print("\n🌍 Creating Geographic Analysis...")

if "Geographic Analysis" in wb.sheetnames:
    geo_sheet = wb["Geographic Analysis"]
    
    regions = []
    region_revenues = []
    
    for row in geo_sheet.iter_rows(min_row=4, max_row=20):
        if row[0].value and row[1].value:
            try:
                region = str(row[0].value).strip()
                revenue = float(row[1].value) if row[1].value else 0
                
                if revenue > 0:
                    regions.append(region)
                    region_revenues.append(revenue)
            except (ValueError, TypeError):
                continue
    
    if regions and region_revenues:
        fig = go.Figure(data=[
            go.Pie(labels=regions, values=region_revenues, hole=0.4)
        ])
        
        fig.update_layout(
            title="Revenue by Geographic Region",
            template='plotly_white',
            height=600
        )
        
        output_file = f"{output_dir}/geographic_breakdown_from_excel.html"
        fig.write_html(output_file)
        print(f"✅ Created: {output_file}")

print("\n" + "="*60)
print("✅ All visualizations created successfully!")
print(f"📂 Location: {output_dir}")
print("="*60)

# List all created files
print("\n📄 Generated Files:")
for file in os.listdir(output_dir):
    if file.endswith('_from_excel.html'):
        print(f"   - {file}")

# Open the first visualization
first_viz = f"{output_dir}/revenue_comparison_from_excel.html"
if os.path.exists(first_viz):
    print(f"\n🌐 Opening: {first_viz}")
    os.system(f'start "" "{first_viz}"')
